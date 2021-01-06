# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *






# globally declare wb and sheet variable 53c92803ebd


# opening the existing excel file 
wb = load_workbook('D:\\stuffregister.xlsx') 



# create the sheet object 
sheet = wb.active 



def excel(): 
	
	# resize the width of columns in 
	# excel spreadsheet 
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	# write given data to an excel spreadsheet 
	# at particular location 
	sheet.cell(row=1, column=1).value = "Name"
	sheet.cell(row=1, column=2).value = "Age"
	sheet.cell(row=1, column=3).value = "Education"
	sheet.cell(row=1, column=4).value = "ID Card Number"
	sheet.cell(row=1, column=5).value = "Contact Number"
	sheet.cell(row=1, column=6).value = "Email id"
	sheet.cell(row=1, column=7).value = "Address"


# Function to set focus (cursor) 
def focus1(event): 
	# set focus on the age_field box 
	age_field.focus_set() 


# Function to set focus 
def focus2(event): 
	# set focus on the education_field box 
	education_field.focus_set() 


# Function to set focus 
def focus3(event): 
	# set focus on the form_no_field box 
	form_no_field.focus_set() 


# Function to set focus 
def focus4(event): 
	# set focus on the contact_no_field box 
	contact_no_field.focus_set() 


# Function to set focus 
def focus5(event): 
	# set focus on the email_id_field box 
	email_id_field.focus_set() 


# Function to set focus 
def focus6(event): 
	# set focus on the address_field box 
	address_field.focus_set() 


# Function for clearing the 
# contents of text entry boxes 
def clear(): 
	
	# clear the content of text entry box 
	name_field.delete(0, END) 
	age_field.delete(0, END) 
	education_field.delete(0, END) 
	IDcard_no_field.delete(0, END) 
	contact_no_field.delete(0, END) 
	email_id_field.delete(0, END) 
	address_field.delete(0, END) 


# Function to take data from GUI 
# window and write to an excel file 
def insert(): 
	
	# if user not fill any entry 
	# then print "empty input" 
	if (name_field.get() == "" and
		age_field.get() == "" and
		education_field.get() == "" and
		IDcard_no_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""): 
			
		print("empty input") 

	else: 

		# assigning the max row and max column 
		# value upto which data is written 
		# in an excel sheet to the variable 
		current_row = sheet.max_row 
		current_column = sheet.max_column 

		# get method returns current text 
		# as string which we write into 
		# excel spreadsheet at particular location 
		sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
		sheet.cell(row=current_row + 1, column=2).value = age_field.get() 
		sheet.cell(row=current_row + 1, column=3).value = education_field.get() 
		sheet.cell(row=current_row + 1, column=4).value = IDcard_no_field.get() 
		sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get() 
		sheet.cell(row=current_row + 1, column=6).value = email_id_field.get() 
		sheet.cell(row=current_row + 1, column=7).value = address_field.get() 

		# save the file 
		wb.save('D:\\stuffregister.xlsx') 

		# set focus on the name_field box 
		name_field.focus_set() 

		# call the clear() function 
		clear() 


# Driver code 
if __name__ == "__main__": 
	
	# create a GUI window 
	root = Tk() 
    
    # disavle resize feature
    root.resizable(False, False)

	# set the background colour of GUI window 
	root.configure(background='light blue') 

	# set the title of GUI window 
	root.title("სარეგისტრაციო ფორმა") 

	# set the configuration of GUI window 
	root.geometry("500x300") 
    
	excel() 

	# create a Form label 
	heading = Label(root, text="სარეგისტრაციო ფორმა", bg="green", fg="white") 

	# create a Name label 
	name = Label(root, text="სახელი", bg="light blue") 

	# create a age label 
	age = Label(root, text="დაბადების თარიღი", bg="light blue") 

	# create a Semester label 
	education = Label(root, text="განათლება", bg="light blue") 

	# create a Form No. lable 
	IDcard_no = Label(root, text="ID ბარათის ნომერი", bg="light blue") 

	# create a Contact No. label 
	contact_no = Label(root, text="საკონტაქტო ტელეფონი", bg="light blue") 

	# create a Email id label 
	email_id = Label(root, text="ელ-ფოსტა", bg="light blue") 

	# create a address label 
	address = Label(root, text="მისამართი", bg="light blue") 

	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
	heading.grid(row=0, column=1) 
	name.grid(row=1, column=0) 
	age.grid(row=2, column=0) 
	education.grid(row=3, column=0) 
	IDcard_no.grid(row=4, column=0) 
	contact_no.grid(row=5, column=0) 
	email_id.grid(row=6, column=0) 
	address.grid(row=7, column=0) 

	# create a text entry box 
	# for typing the information 
	name_field = Entry(root) 
	age_field = Entry(root) 
	education_field = Entry(root) 
	IDcard_no_field = Entry(root) 
	contact_no_field = Entry(root) 
	email_id_field = Entry(root) 
	address_field = Entry(root) 

	# bind method of widget is used for 
	# the binding the function with the events 

	# whenever the enter key is pressed 
	# then call the focus1 function 
	name_field.bind("<Return>", focus1) 

	# whenever the enter key is pressed 
	# then call the focus2 function 
	age_field.bind("<Return>", focus2) 

	# whenever the enter key is pressed 
	# then call the focus3 function 
	education_field.bind("<Return>", focus3) 

	# whenever the enter key is pressed 
	# then call the focus4 function 
	IDcard_no_field.bind("<Return>", focus4) 

	# whenever the enter key is pressed 
	# then call the focus5 function 
	contact_no_field.bind("<Return>", focus5) 

	# whenever the enter key is pressed 
	# then call the focus6 function 
	email_id_field.bind("<Return>", focus6) 

	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
	name_field.grid(row=1, column=1, ipadx="100") 
	age_field.grid(row=2, column=1, ipadx="100") 
	education_field.grid(row=3, column=1, ipadx="100") 
	IDcard_no_field.grid(row=4, column=1, ipadx="100") 
	contact_no_field.grid(row=5, column=1, ipadx="100") 
	email_id_field.grid(row=6, column=1, ipadx="100") 
	address_field.grid(row=7, column=1, ipadx="100") 

	# call excel function 
	excel() 

	# create a Submit Button and place into the root window 
	submit = Button(root, text="დადასტურება", fg="white", 
							bg="Green", command=insert) 
	submit.grid(row=8, column=1) 

	exit = Button(root, text="დახურვა",  fg="white", bg="Green", command=root.quit  )
	exit.grid(row=9, column=1) 
	


	root.wm_iconbitmap('D:/pythonExamples/Pregistration/53c92803ebd.ico')

	# start the GUI 
	root.mainloop() 
